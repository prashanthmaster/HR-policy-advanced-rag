"""
Spreadsheet parser -- XLSX and CSV share this module because, once past
the file-format-specific loading step, they're the same problem: a header
row plus data rows that need to become one row-wise retrieval piece each
(mirroring ingestion.table_serializer's TableRowPiece for the markdown
housing table). This is the concrete case for building it once: an Excel
export and a CSV export of the exact same table should turn into the
exact same retrieval pieces, and sharing the code is what guarantees that
rather than hoping two separate implementations stay in sync.

Manifest locator for a spreadsheet table: {"sheet_name": str (XLSX only),
"header_row": int, "row_start": int, "row_end": int} -- all 1-indexed
inclusive, matching how a person reading the file in Excel would describe
"header in row 1, data rows 2 through 14."

The header row's cell text becomes each row's column labels, joined into
the row's text as "<header>: <value>" pairs -- so a row from this Excel
file reads the same way as the corresponding markdown bullet row already
does ("Grade M2, 3 to 5 years of service: 9,500 per month.") is NOT
reproduced verbatim -- this module produces its own row text shape,
labelled by column header, since a spreadsheet has no natural prose
sentence to copy the way the markdown corpus's author wrote it. Tested
independently against its own expected shape, not against the markdown
table's exact wording.
"""

from __future__ import annotations

import csv as csv_module
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

from ingestion.formats.manifest import load_manifest
from ingestion.logging_setup import get_logger

_log = get_logger("ingestion.formats.spreadsheet_parser")


@dataclass
class SpreadsheetRowPiece:
    clause_id: str
    piece_id: str
    row_index: int
    total_rows: int
    text: str
    metadata_fields: dict


def _row_text(headers: list[str], row_values: list[object], title: str | None) -> str:
    pairs = [f"{h.strip()}: {v}" for h, v in zip(headers, row_values) if h and v not in (None, "")]
    parts = [title] if title else []
    parts.append(", ".join(pairs) + ".")
    return " ".join(p for p in parts if p)


def _rows_from_entry(all_rows: list[list[object]], locator: dict) -> tuple[list[str], list[list[object]]]:
    header_row = locator["header_row"]
    row_start = locator["row_start"]
    row_end = locator["row_end"]
    headers = [str(c) if c is not None else "" for c in all_rows[header_row - 1]]
    data = all_rows[row_start - 1 : row_end]
    return headers, data


def parse_xlsx(xlsx_path: Path, manifest_path: Path, *, repo_root: Path | None = None) -> list[SpreadsheetRowPiece]:
    wb = load_workbook(str(xlsx_path), data_only=True)
    entries = load_manifest(manifest_path)

    pieces: list[SpreadsheetRowPiece] = []
    for entry in entries:
        locator = entry["locator"]
        ws = wb[locator["sheet_name"]] if "sheet_name" in locator else wb.active
        all_rows = [list(row) for row in ws.iter_rows(values_only=True)]
        headers, data = _rows_from_entry(all_rows, locator)
        title = entry.get("title")

        metadata_fields = {k: v for k, v in entry.items() if k not in ("locator", "title")}
        total = len(data)
        for i, row in enumerate(data):
            pieces.append(
                SpreadsheetRowPiece(
                    clause_id=entry["clause_id"],
                    piece_id=f"{entry['clause_id']}#row{i}",
                    row_index=i,
                    total_rows=total,
                    text=_row_text(headers, row, title),
                    metadata_fields=metadata_fields,
                )
            )
    _log.info("parsed %d row pieces from XLSX %s", len(pieces), xlsx_path)
    return pieces


def parse_csv(csv_path: Path, manifest_path: Path, *, repo_root: Path | None = None) -> list[SpreadsheetRowPiece]:
    with open(csv_path, newline="", encoding="utf-8") as f:
        all_rows = [list(row) for row in csv_module.reader(f)]
    entries = load_manifest(manifest_path)

    pieces: list[SpreadsheetRowPiece] = []
    for entry in entries:
        locator = entry["locator"]
        headers, data = _rows_from_entry(all_rows, locator)
        title = entry.get("title")

        metadata_fields = {k: v for k, v in entry.items() if k not in ("locator", "title")}
        total = len(data)
        for i, row in enumerate(data):
            pieces.append(
                SpreadsheetRowPiece(
                    clause_id=entry["clause_id"],
                    piece_id=f"{entry['clause_id']}#row{i}",
                    row_index=i,
                    total_rows=total,
                    text=_row_text(headers, row, title),
                    metadata_fields=metadata_fields,
                )
            )
    _log.info("parsed %d row pieces from CSV %s", len(pieces), csv_path)
    return pieces
